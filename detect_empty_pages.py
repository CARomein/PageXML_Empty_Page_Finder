#!/usr/bin/env python3
"""
Empty Page Detection Tool for Transkribus PageXML
Scans PageXML collections and identifies pages without transcribed text content.
"""

import xml.etree.ElementTree as ET
from pathlib import Path
import argparse
import sys
from typing import List, Tuple


class EmptyPageDetector:
    """Detects pages without text content in PageXML collections."""
    
    def __init__(self, base_path: Path, quiet: bool = False):
        self.base_path = Path(base_path)
        self.quiet = quiet
        self.namespace = {'page': 'http://schema.primaresearch.org/PAGE/gts/pagecontent/2013-07-15'}
        self.empty_pages: List[Tuple[str, str, str]] = []
        
    def log(self, message: str, end: str = '\n'):
        """Print message unless in quiet mode."""
        if not self.quiet:
            print(message, end=end, flush=True)
    
    def find_collections(self) -> List[Path]:
        """Discover all collection directories containing page/ subdirectories."""
        collections = []
        
        if not self.base_path.exists():
            raise FileNotFoundError(f"Base path does not exist: {self.base_path}")
        
        for item in sorted(self.base_path.iterdir()):
            if item.is_dir():
                page_dir = item / 'page'
                if page_dir.exists() and page_dir.is_dir():
                    xml_files = list(page_dir.glob('*.xml'))
                    if xml_files:
                        collections.append(item)
        
        return collections
    
    def is_page_empty(self, xml_path: Path) -> bool:
        """
        Check if a PageXML file contains any transcribed text.
        Returns True if the page has no text content.
        """
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            # Find all TextLine elements
            text_lines = root.findall('.//page:TextLine', self.namespace)
            
            if not text_lines:
                # No text lines at all means empty page
                return True
            
            # Check if any TextLine contains non-empty Unicode text
            for text_line in text_lines:
                unicode_elem = text_line.find('.//page:Unicode', self.namespace)
                if unicode_elem is not None and unicode_elem.text:
                    text_content = unicode_elem.text.strip()
                    if text_content:
                        # Found non-empty text, page is not empty
                        return False
            
            # All TextLines were empty or had no Unicode elements
            return True
            
        except ET.ParseError as e:
            self.log(f"  Warning: Could not parse {xml_path.name}: {e}")
            return False
        except Exception as e:
            self.log(f"  Warning: Error processing {xml_path.name}: {e}")
            return False
    
    def get_image_filename(self, xml_path: Path) -> str:
        """Extract the image filename from PageXML metadata."""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            page_elem = root.find('.//page:Page', self.namespace)
            
            if page_elem is not None:
                image_filename = page_elem.get('imageFilename')
                if image_filename:
                    return image_filename
            
            # Fallback to XML filename if imageFilename not found
            return xml_path.stem
            
        except Exception:
            return xml_path.stem
    
    def process_collection(self, collection_path: Path):
        """Process all PageXML files in a collection."""
        collection_name = collection_path.name
        page_dir = collection_path / 'page'
        xml_files = sorted(page_dir.glob('*.xml'))
        
        self.log(f"\n  Processing collection: {collection_name}")
        self.log(f"  Found {len(xml_files)} XML files")
        
        empty_count = 0
        for i, xml_file in enumerate(xml_files, 1):
            if not self.quiet and i % 10 == 0:
                self.log(f"    Processed {i}/{len(xml_files)} files", end='\r')
            
            if self.is_page_empty(xml_file):
                image_filename = self.get_image_filename(xml_file)
                self.empty_pages.append((collection_name, image_filename, xml_file.name))
                empty_count += 1
        
        if not self.quiet:
            self.log(f"    Processed {len(xml_files)}/{len(xml_files)} files    ")
        self.log(f"  Found {empty_count} empty page(s) in {collection_name}")
    
    def run(self) -> List[Tuple[str, str, str]]:
        """Execute the empty page detection process."""
        self.log("=== Empty Page Detection Tool ===\n")
        
        collections = self.find_collections()
        
        if not collections:
            self.log("No collections found. Please check the directory structure.")
            self.log("Expected structure: base_path/Collection_Name/page/*.xml")
            return []
        
        self.log(f"Found {len(collections)} collection(s):")
        for collection in collections:
            self.log(f"  - {collection.name}")
        
        for collection in collections:
            self.process_collection(collection)
        
        self.log(f"\n=== Summary ===")
        self.log(f"Total empty pages found: {len(self.empty_pages)}")
        self.log(f"Collections processed: {len(collections)}")
        
        return self.empty_pages


def write_to_excel(empty_pages: List[Tuple[str, str, str]], output_path: Path):
    """Write empty pages data to Excel file using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty Pages"
        
        # Create header row
        headers = ['Collection', 'Image Filename', 'XML Filename']
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for collection, image_filename, xml_filename in empty_pages:
            ws.append([collection, image_filename, xml_filename])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        
        # Save workbook
        wb.save(output_path)
        return True
        
    except ImportError:
        return False


def write_to_csv_fallback(empty_pages: List[Tuple[str, str, str]], output_path: Path):
    """Fallback: Write to CSV if openpyxl is not available."""
    import csv
    
    csv_path = output_path.with_suffix('.csv')
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Collection', 'Image Filename', 'XML Filename'])
        writer.writerows(empty_pages)
    
    return csv_path


def main():
    """Main execution function."""
    parser = argparse.ArgumentParser(
        description='Detect pages without transcribed text in Transkribus PageXML collections',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument(
        'base_path',
        nargs='?',
        help='Path to base directory containing PageXML collections'
    )
    
    parser.add_argument(
        '-o', '--output',
        help='Output Excel file path (default: empty_pages.xlsx in script directory)',
        default=None
    )
    
    parser.add_argument(
        '-q', '--quiet',
        action='store_true',
        help='Suppress progress output'
    )
    
    args = parser.parse_args()
    
    # Get base path from argument or prompt
    if args.base_path:
        base_path = Path(args.base_path)
    else:
        base_path_input = input("Enter path to collections directory: ").strip()
        if not base_path_input:
            print("Error: No path provided")
            sys.exit(1)
        base_path = Path(base_path_input)
    
    # Set output path
    if args.output:
        output_path = Path(args.output)
    else:
        # Default: place in script directory
        script_dir = Path(__file__).parent.resolve()
        output_path = script_dir / 'empty_pages.xlsx'
    
    try:
        # Run detection
        detector = EmptyPageDetector(base_path, args.quiet)
        empty_pages = detector.run()
        
        if not empty_pages:
            print("\n✓ No empty pages found. All pages contain transcribed text.")
            return
        
        # Write output
        print(f"\nGenerating output file...")
        
        excel_success = write_to_excel(empty_pages, output_path)
        
        if excel_success:
            print(f"✓ Report generated: {output_path}")
            print(f"  Open the file in Excel or LibreOffice to view the results")
        else:
            # Fallback to CSV
            print("Note: openpyxl not available, creating CSV instead")
            csv_path = write_to_csv_fallback(empty_pages, output_path)
            print(f"✓ Report generated: {csv_path}")
            print(f"  Open the file in Excel or a text editor to view the results")
            print(f"\nTo enable Excel output, install openpyxl:")
            print(f"  pip install openpyxl --break-system-packages")
    
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except KeyboardInterrupt:
        print("\n\nProcess interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
